# TAG 팀 프로젝트 python 코딩 by  김재웅 &  권준서
import openpyxl
import os
import getpass

from time import time
from time import sleep
from time import strftime
from time import localtime

def del_ref():                              # delay&refresh : delay 0.5초 후 화면 새로고침
    sleep(0.5)
    os.system('cls')

password = '1q2w3e4r'                       # 관리자 비밀번호

excel_name = 'TAGDATA.xlsx'
excel_log = 'TAGLOG.xlsx'
wb = openpyxl.load_workbook(filename=excel_name)         # filename= 에 엑셀 data 파일, wb: 엑셀 파일 자체
wb2 = openpyxl.load_workbook(filename=excel_log)

WS_H = wb['사람']                                        # WS_H : 사용자 DB, WS_TL : 공구 DB, WS_NOW : 대여현황 DB, WS_ACCESSLIST : 출입명부
WS_TL = wb['공구']
WS_3DLIST = wb['3D프린터명단']
WS_LOG = wb2['기록']
WS_ACCESSLIST = wb2['출입명부']
WS_3D = wb2['3D프린터이용대장']

def LIST3D():
    printer_list = []
    LIST3D_RANGE = range(WS_3DLIST_COUNTER-3)
    for i in LIST3D_RANGE:
        printer_list.append(WS_3DLIST.cell(row=i+2, column=1).value)


WS_H_COUNTER= WS_H.max_row + 1                           # DB 마지막 줄 COUNTER 등록
WS_TL_COUNTER = WS_TL.max_row + 1
WS_LOG_COUNTER = WS_LOG.max_row + 1
WS_ACCESSLIST_COUNTER = WS_ACCESSLIST.max_row + 1
WS_3D_COUNTER = WS_3D.max_row + 1
WS_3DLIST_COUNTER = WS_3DLIST.max_row + 1

NewH_counter = 0

def findUID(WS,WS_COUNTER,col,UID):         # UID 검색 함수(WS: WS 종류, WS_COUNTER: WS COUNTER 종류, col: 행, UID: 검색 UID)
    Rang = range(2, WS_COUNTER)          # 함수 결과는 UID가 있으면 UID가 몇번째 열에 있는지 나옴
    location = 2                         # UID가 DB에 없다면 -1 출력
    if WS_COUNTER == 2:
        return -1
    else:
        for i in Rang:
            if i < WS_COUNTER-1:
                if UID == WS.cell(row=location,column=col).value:
                    return location
                else:
                    location = location+1
                    continue
            elif i == WS_COUNTER-1:
                if UID == WS.cell(row=location,column=col).value:
                    return location
                else:
                    return -1

def CALL_H(WS_H_NUMBER):                    # 사용자 DB에서 WS_H_NUMBER 행의 정보를 불러오는 함수
    H_UID = WS_H.cell(row=WS_H_NUMBER, column=1).value
    H_NAME = WS_H.cell(row=WS_H_NUMBER, column=2).value
    H_ID = WS_H.cell(row=WS_H_NUMBER, column=3).value
    H_PN = WS_H.cell(row=WS_H_NUMBER, column=4).value
    H_depertment = WS_H.cell(row=WS_H_NUMBER, column=5).value
    print(H_UID," ",H_NAME," ",H_depertment," ",H_ID," ",H_PN,"\n")

def CALL_TL(WS_TL_NUMBER):                  # 수공구 DB에서 WS_H_NUMBER 행의 정보를 불러오는 함수
    TL_UID = WS_TL.cell(row=WS_TL_NUMBER, column=1).value
    TL_TYPE = WS_TL.cell(row=WS_TL_NUMBER, column=2).value
    TL_NAME = WS_TL.cell(row=WS_TL_NUMBER, column=3).value
    TL_OX = WS_TL.cell(row=WS_TL_NUMBER, column=4).value
    TL_RENTAIL_NAME = WS_TL.cell(row=WS_TL_NUMBER, column=5).value
    print(TL_UID," ",TL_TYPE," ",TL_NAME," ",TL_OX," ",TL_RENTAIL_NAME,"\n")

def CALL_TIME():                             # 로그 시간 기록
    tm = localtime(time())
    return strftime('%Y-%m-%d %I:%M:%S %p', tm)

def CALL_TIME2():
    tm = localtime(time())
    return strftime('%Y-%m-%d',tm)


while 1:                                                                # 무한 반복 루프
    os.system('cls')
    print("Welcome to MAKERSPACE KU\n")
    print("1. 출입명부 작성")
    print("2. 3D프린터 이용대장 작성")
    print("3. 수공구 대여/반납/대여확인")
    print("4. 신규등록")
    print("5. 관리자 모드\n")
    
    
    choice = input("")
    # Main - 1. 출입 명부 작성 모드
    if choice == '1':
        os.system('cls')
        print("출입 명부 작성 모드\n")
        print("학생증을 태그해주십시오")
        inputuid = input("")
        H_location = findUID(WS_H,WS_H_COUNTER,1,inputuid)
        os.system('cls')
        if H_location == -1:
            print("등록되지 않은 사용자입니다. 신규등록을 해주십시오")
            sleep(0.5)
            del_ref()
        else:
            while 1:
                H_name = WS_H.cell(row=H_location, column=2).value
                print(H_name,"님 환영합니다.\n\n방문목적에 해당하는 번호를 입력해주십시오")
                print("1.테이블 사용\n2.유저 동행\n3.기타\n4.취소")
                choice_2 = input("") 
                if choice_2 == '1':
                    os.system('cls')
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=1,value=WS_ACCESSLIST_COUNTER-1)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=2,value=CALL_TIME())
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=3,value=H_name)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=4,value=WS_H.cell(row=H_location, column=5).value)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=5,value=WS_H.cell(row=H_location, column=4).value)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=6,value="테이블 사용")
                    wb2.save(excel_log)
                    WS_ACCESSLIST_COUNTER += 1
                    print(H_name,"님 테이블 사용 목적으로 방문하셨습니다")
                    sleep(0.3)
                    del_ref()
                    break
                elif choice_2 == '2':
                    os.system('cls')
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=1,value=WS_ACCESSLIST_COUNTER-1)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=2,value=CALL_TIME())
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=3,value=H_name)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=4,value=WS_H.cell(row=H_location, column=5).value)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=5,value=WS_H.cell(row=H_location, column=4).value)
                    WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=6,value="유저동행")
                    wb2.save(excel_log)
                    WS_ACCESSLIST_COUNTER += 1
                    print(H_name,"님 유저 동행 목적으로 방문하셨습니다")
                    sleep(0.3)
                    del_ref()
                    break
                elif choice_2 == '3':
                    os.system('cls')
                    print("기타 방문 목적을 입력해주십시오")
                    input_purpose = input("")
                    os.system('cls')
                    print(H_name,"님 방문목적:",input_purpose,"\n맞으면 1 아니면 1을 제외한 아무키나 입력해주십시오")
                    inputOX = input("")
                    if inputOX == '1':
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=1,value=WS_ACCESSLIST_COUNTER-1)
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=2,value=CALL_TIME())
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=3,value=H_name)
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=4,value=WS_H.cell(row=H_location, column=5).value)
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=5,value=WS_H.cell(row=H_location, column=4).value)
                        WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER,column=6,value=input_purpose)
                        wb2.save(excel_log)
                        WS_ACCESSLIST_COUNTER += 1
                        print(H_name,"님",input_purpose,"목적으로 방문하셨습니다")
                        sleep(0.3)
                        del_ref()
                        break
                    else:
                        print("취소되었습니다")
                        sleep(0.3)
                        del_ref()                        
                elif choice_2 == '4':
                    os.system('cls')
                    print("출입명부 작성이 취소되었습니다")
                    sleep(0.3)
                    del_ref()
                    break
                else:
                    os.system('cls')
                    print("잘못된 입력입니다")
                    sleep(0.3)
                    del_ref()

    elif choice == '2':
        os.system('cls')
        print("3D프린터 이용대장 작성 모드\n")
        print("학생증을 태그해주십시오")
        inputuid = input("")
        H_location = findUID(WS_H,WS_H_COUNTER,1,inputuid)
        os.system('cls')
        if H_location == -1:
            print("등록되지 않은 사용자입니다. 신규등록을 해주십시오")
            sleep(0.5)
            del_ref()
        else:
            H_name = WS_H.cell(row=H_location, column=2).value
            print(H_name,"님 환영합니다.\n\n3D 프린터명을 입력해주십시오")
            printer_3D = input("")
            if len(printer_3D) <4 :
                printer_3D = printer_3D.upper()
                os.system('cls')
                print(printer_3D,"를 선택했습니다\n이용 시작시간을 입력해주십시오(ex.13:00)")
                printer_3D_start_time = input("")
                print("이용 종료시간을 입력해주십시오(ex.15:00)")
                printer_3D_end_time = input("")
                printer_3d_time = printer_3D_start_time + "-" + printer_3D_end_time
                os.system('cls')
                print("날짜: ",CALL_TIME2())
                print("이용자명: ",H_name)
                print("프린터명: ",printer_3D)
                print("사용시간: ",printer_3d_time)
                print("입력한 정보가 맞으면 1, 틀리거나 취소하려면 1을 제외한 키를 입력해주십시오")
                inputOX = input("")
                if inputOX == '1':
                    os.system('cls')
                    WS_3D.cell(row=WS_3D_COUNTER,column=1,value=CALL_TIME2())
                    WS_3D.cell(row=WS_3D_COUNTER,column=2,value=WS_H.cell(row=H_location, column=5).value)
                    WS_3D.cell(row=WS_3D_COUNTER,column=3,value=WS_H.cell(row=H_location, column=3).value)
                    WS_3D.cell(row=WS_3D_COUNTER,column=4,value=H_name)
                    WS_3D.cell(row=WS_3D_COUNTER,column=5,value=printer_3D)
                    WS_3D.cell(row=WS_3D_COUNTER,column=6,value=printer_3d_time)
                    wb2.save(excel_log)
                    print("3D프린터 사용이 등록되었습니다")
                    sleep(0.3)
                    del_ref()
                else:
                    os.system('cls')
                    print("취소되었습니다")
                    sleep(0.3)
                    del_ref()
            else: 
                print("잘못된 입력입니다")
                del_ref()


    # Main - 3. 대여/반납/대여 수공구 확인 모드
    elif choice == '3':
        while 1:
            os.system('cls')
            print("수공구 대여/반납/대여확인 모드\n")        # WS_TL(수공구_TOOL DB) 접근
            print("1. 대여\n2. 반납\n3. 대여 수공구 확인\n4. 나가기")
            choice_1 = input("")
            # 1-1. 대여 모드
            if choice_1 == '1':
                os.system('cls')
                print("대여 모드\n")
                print("학생증을 태그해주십시오")
                rental_uid = input("")
                H_location = findUID(WS_H,WS_H_COUNTER,1,rental_uid)
                if H_location == -1:
                    os.system('cls')
                    print("등록되지 않은 사용자입니다. 신규등록을 해주십시오")
                    sleep(0.3)
                    del_ref()
                    break
                else:
                    os.system('cls')
                    rental_name = WS_H.cell(row=H_location,column=2).value
                    print(rental_name,"님 환영합니다\n")
                    while_exit = 0
                    rental_list = []
                    while while_exit == 0:
                        print("수공구 RFID를 태그해주십시오. 대여가 완료되면 '1'을 입력해주십이오")
                        inputuid = input("")
                        if inputuid == '1':
                            if len(rental_list) == 0:
                                os.system('cls')
                                print("대여한 수공구가 없습니다.")
                                sleep(0.3)
                                del_ref()
                                break
                            else:
                                os.system('cls')
                                rental_length = len(rental_list)
                                rental_range = range(rental_length)                                
                                rental_list_name = []
                                for i in rental_range:
                                    print(WS_TL.cell(row=rental_list[i],column=3).value)
                                    rental_list_name.append(WS_TL.cell(row=rental_list[i],column=3).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=rental_uid)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=rental_name)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="대여")
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=''.join(s for s in rental_list_name))
                                wb.save(excel_name)
                                wb2.save(excel_log)
                                WS_LOG_COUNTER = WS_LOG_COUNTER + 1
                                print("의 대여가 완료되었습니다")
                                rental_list=[]
                                sleep(0.7)
                                del_ref()
                                break
                            break
                        else:
                            os.system('cls')
                            TL_location = findUID(WS_TL,WS_TL_COUNTER,1,inputuid)
                            if TL_location == -1:
                                print("등록되지 않은 수공구입니다. 다시 확인해주십시오")
                                sleep(0.3)
                                del_ref()
                            else:
                                if WS_TL.cell(row=TL_location,column=4).value == 1:
                                    print("이미 대여된 수공구입니다")
                                    del_ref()
                                elif WS_TL.cell(row=TL_location,column=4).value == 0:
                                    WS_TL.cell(row=TL_location,column=4,value=1)
                                    WS_TL.cell(row=TL_location,column=5,value=rental_uid)                                    
                                    print(WS_TL.cell(row=TL_location,column=3).value,"를 대여했습니다")
                                    rental_list.append(TL_location)
                                    del_ref()
            # 1-2. 반납 모드
            elif choice_1 == '2':
                os.system('cls')
                print("반납 모드\n")
                print("학생증을 태그해주십시오")
                inputuid = input("")
                return_uid = inputuid
                H_location = findUID(WS_H,WS_H_COUNTER,1,inputuid)
                if H_location == -1:
                    os.system('cls')
                    print("등록되지 않은 사용자입니다. 신규등록을 해주십시오")
                    sleep(0.3)
                    del_ref()
                    break
                else:
                    os.system('cls')
                    return_name = WS_H.cell(row=H_location,column=2).value
                    return_list = []
                    while 1:
                        print(return_name,"님 환영합니다.\n")
                        TL_rental_list = []
                        TL_rental_range = range(2,WS_TL_COUNTER)
                        if WS_TL_COUNTER == 2:
                            print("대여한 수공구가 없습니다\n")                            
                        else:
                            for i in TL_rental_range:
                                if i <= WS_TL_COUNTER-1:
                                    if return_uid == WS_TL.cell(row=i,column=5).value:
                                        TL_rental_list.append(i)
                            if len(TL_rental_list) == 0:
                                print("대여한 수공구가 없습니다\n")
                            else:
                                print("대여하신 수공구는")
                                TL_rental_list_range = range(len(TL_rental_list))
                                for i in TL_rental_list_range:
                                    print(WS_TL.cell(row=TL_rental_list[i],column=3).value)
                                print("\n입니다\n")

                        
                        print("반납하실 수공구 RFID를 태그해주십시오. 반납이 완료되면 '1'을 입력해주십이오")
                        inputuid = input("")
                        if inputuid == '1':
                            if len(return_list) == 0:
                                os.system('cls')
                                print("반납한 수공구가 없습니다")
                                sleep(0.3)
                                del_ref()
                                break
                            else:
                                os.system('cls')
                                return_length = len(return_list)
                                return_range = range(return_length)
                                return_list_name = []
                                for i in return_range:
                                    print(WS_TL.cell(row=return_list[i],column=3).value)
                                    return_list_name.append(WS_TL.cell(row=return_list[i],column=3).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=return_uid)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=return_name)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="반납")
                                WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=''.join(s for s in return_list_name))
                                wb.save(excel_name)
                                wb2.save(excel_log)
                                print("의 반납이 완료되었습니다")
                                WS_LOG_COUNTER = WS_LOG_COUNTER + 1
                                return_list=[]
                                sleep(0.7)
                                del_ref()
                                break
                            
                        else:
                            TL_location = findUID(WS_TL,WS_TL_COUNTER,1,inputuid)
                            if TL_location == -1:
                                print("등록되지 않은 수공구입니다. 다시 확인해주십시오")
                                sleep(0.3)
                                del_ref()
                            else:
                                if WS_TL.cell(row=TL_location,column=4).value == 0:
                                    print("대여하지 않은 수공구입니다. 다시 확인해주십시오")
                                    del_ref()
                                elif WS_TL.cell(row=TL_location,column=4).value == 1:
                                    if return_uid == WS_TL.cell(row=TL_location,column=5).value:
                                        WS_TL.cell(row=TL_location,column=4,value=0)
                                        WS_TL.cell(row=TL_location,column=5,value="")
                                        print(WS_TL.cell(row=TL_location,column=3).value,"를 반납했습니다")
                                        return_list.append(TL_location)
                                        del_ref()
                                    else:
                                        print("대여한 사용자가 다른 수공구입니다")
                                        del_ref()  
                    break
            # 1-3. 대여 현황 확인    
            elif choice_1 == '3':
                os.system('cls')
                print("대여 수공구 확인 모드\n")
                print("학생증을 태그해주십시오")
                inputuid = input("")
                H_location = findUID(WS_H,WS_H_COUNTER,1,inputuid)
                H_name = WS_H.cell(row=H_location, column=2).value
                os.system('cls')
                if H_location == -1:
                    print("등록되지 않은 사용자입니다. 신규등록을 해주십시오")
                    sleep(0.3)
                    del_ref()
                else:
                    print(H_name,"님 환영합니다.\n\n대여하신 수공구는\n")
                    TL_rental_list = []
                    TL_rental_range = range(2,WS_TL_COUNTER)
                    if WS_TL_COUNTER == 2:
                        print("대여한 수공구가 없습니다")
                        sleep(0.3)
                        del_ref()
                    else:
                        for i in TL_rental_range:
                            if i <= WS_TL_COUNTER-1:
                                if inputuid == WS_TL.cell(row=i,column=5).value:
                                    TL_rental_list.append(i)
                        if len(TL_rental_list) == 0:
                            print("대여한 수공구가 없습니다")
                            sleep(0.3)
                            del_ref()
                        else:
                            TL_rental_list_range = range(len(TL_rental_list))
                            for i in TL_rental_list_range:
                                print(WS_TL.cell(row=TL_rental_list[i],column=3).value)
                            print("\n입니다\n\n확인을 완료했으면 아무키나 입력해주십시오")
                            inputkey = input("")
            # 1-4. 나가기
            elif choice_1 == '4':
                os.system('cls')
                break
            # 1-else. 잘못된 입력의 경우
            else:
                print("잘못된 입력입니다!")
                del_ref()
        


                      
    # Main - 4. 신규 등록 모드
    elif choice == '4':
        os.system('cls')
        print("신규등록 모드\n")         # WS_H(사람_HUMAN DB) 접근
        print("학생증을 태그해주십시오")
        inputuid = input("")
        NewH_location = findUID(WS_H,WS_H_COUNTER,1,inputuid)
        if not NewH_location == -1:
                os.system('cls')
                print("이미 등록된 사용자입니다.")
                sleep(0.8)
                del_ref()
        else:
            del_ref()
            print("신규등록 모드\n") 
            print("이름을 입력해주십시오")
            inputname = input("")
            os.system('cls')
            print("신규등록 모드\n")
            print("학과를 입력해주십시오")
            inputdepertment = input("")
            os.system('cls')
            print("신규등록 모드\n") 
            print("학번을 입력해주십시오")
            inputID = input("")
            os.system('cls')
            print("신규등록 모드\n") 
            print("전화번호를 입력해주십시오(-는 생략)")
            inputPN = input("")
            os.system('cls')
            print("신규등록 모드\n") 
            print("학생증 UID: " , inputuid)
            print("이름: " , inputname)
            print("학과: ", inputdepertment)
            print("학번: " , inputID)
            print("전화번호: " , inputPN)
            print("입력한 정보가 맞으면 1, 아니면 2를 입력해주십시오\n")       # 입력 확인
            choice_3 = input("")
            if choice_3 == '1':                    # 등록
                os.system('cls')
                WS_H.cell(row=WS_H_COUNTER,column=1,value=inputuid)
                WS_H.cell(row=WS_H_COUNTER,column=2,value=inputname)
                WS_H.cell(row=WS_H_COUNTER,column=3,value=inputID)
                WS_H.cell(row=WS_H_COUNTER,column=4,value=inputPN)
                WS_H.cell(row=WS_H_COUNTER,column=5,value=inputdepertment)
                wb.save(excel_name)
                WS_H_COUNTER=WS_H_COUNTER+1
                print("등록이 완료되었습니다\n")
                NewH_counter = NewH_counter + 1
                sleep(0.5)
                del_ref()
            elif not choice_3 == '1':                   #등록 취소
                os.system('cls')
                print("등록이 취소되었습니다\n")
                sleep(0.5)
                del_ref()

    # Main - 5. 관리자 모드
    elif choice == '5' :
        print("관리자 모드 선택")           # xl_TL(수공구_TOOL DB), xl_H(사람_HUMAN DB) 모두 접근 가능
        del_ref()
        
        # 3-0. 관리자 비밀번호 확인 
        auth = 0
        for i in range(1,4):
            print("비밀번호를 입력해주십시오")
            inputpw = getpass.getpass('PW: ')           # 비밀번호 입력시 안보이게 함
            if inputpw == password:
                print("승인되었습니다")
                del_ref()
                auth +=1
                break
            else:
                print("잘못된 비밀번호입니다! 남은 횟수 : ",3-i)
                sleep(1+i)
                os.system('cls')

        if auth == 1:
            while 1:
                os.system('cls')
                print("관리자 모드")
                print("1. 수공구 신규 등록\n2. 대여 수공구 현황 확인 및 관리\n3. 신규 등록 인원 확인\n4. 나가기")
                admin = input("")

                # 3-1. 관리자 모드 - 공구 신규 등록
                if admin == '1':
                    os.system('cls')
                    print("수공구 신규 등록\n")
                    print("수공구 RFID 태그해주십시오")
                    inputuid = input("")
                    del_ref()
                    print("수공구 신규 등록\n")
                    print("수공구 종류를 입력해주십시오")
                    inputtype = input("")
                    while 1:
                        os.system('cls')
                        print("수공구 신규 등록\n")
                        print("수공구 이름을 입력해주십시오")
                        inputname = input("")
                        if findUID(WS_TL,WS_TL_COUNTER,3,inputname) == -1:
                            break
                        else:
                            os.system('cls')
                            print(inputname,"은 이미 등록된 수공구 이름입니다. 다시 확인해주십시오")
                            sleep(0.7)
                            del_ref()
                    os.system('cls')
                    print("수공구 신규 등록\n")
                    print("UID: " , inputuid)
                    print("종류: " , inputtype)
                    print("수공구명: ", inputname)
                    print("입력한 정보가 맞다면 1, 아니면 2를 입력해주십시오")
                    choice_4_1 = input("")
                    if choice_4_1 == '1':
                        NewTL_location = findUID(WS_TL,WS_TL_COUNTER,1,inputuid)
                        if not NewTL_location == -1:
                            os.system('cls')
                            print("이미 등록된 수공구입니다.")
                            sleep(0.8)
                            del_ref()
                        else:
                            os.system('cls')
                            WS_TL.cell(row=WS_TL_COUNTER,column=1,value=inputuid)       # 대여현황 1 : true, 0 : false
                            WS_TL.cell(row=WS_TL_COUNTER,column=2,value=inputtype)
                            WS_TL.cell(row=WS_TL_COUNTER,column=3,value=inputname)
                            WS_TL.cell(row=WS_TL_COUNTER,column=4,value=0)
                            WS_TL.cell(row=WS_TL_COUNTER,column=5,value="")
                            wb.save(excel_name)
                            WS_TL_COUNTER=WS_TL_COUNTER+1
                            print("등록이 완료되었습니다\n")
                            sleep(0.5)
                            del_ref()
                    elif not choice_4_1 == '1':
                        print("등록이 취소되었습니다")
                        sleep(0.5)
                        del_ref()

                # 3-2. 관리자 모드 - 2항
                elif admin == '2':
                    os.system('cls')
                    print("대여 수공구 현황 확인 및 관리\n")
                    TOTAL_rental_list = []
                    TL_rental_range = range(2,WS_TL_COUNTER)
                    if WS_TL_COUNTER == 2:
                        print("대여한 수공구가 없습니다")
                        sleep(0.3)
                        del_ref()
                    else:
                        while 1:
                            for i in TL_rental_range:
                                if i <= WS_TL_COUNTER-1:
                                    if WS_TL.cell(row=i,column=4).value == 1:
                                        TOTAL_rental_list.append(i)
                            if len(TOTAL_rental_list) == 0:
                                print("대여한 수공구가 없습니다")
                                sleep(0.3)
                                del_ref()
                                break
                            else:
                                TL_rental_list_range = range(len(TOTAL_rental_list))
                                for i in TL_rental_list_range:
                                    TOTAL_rental_uid = WS_TL.cell(row=TOTAL_rental_list[i],column=5).value
                                    TOTAL_rental_name = WS_H.cell(row=findUID(WS_H,WS_H_COUNTER,1,TOTAL_rental_uid), column=2).value
                                    print(i+1,".",WS_TL.cell(row=TOTAL_rental_list[i],column=3).value,"-",TOTAL_rental_name)
                                print("\n반납처리할 수공구에 해당하는 번호를 입력해주십시오.\n작업을 완료했으면 0을 입력해주십시오")
                                inputkey = input("")
                                if not inputkey.isdigit():
                                    os.system('cls')
                                    print("잘못된 입력입니다.")
                                    sleep(0.3)
                                    del_ref()
                                else:
                                    inputkey = int(inputkey)
                                    if inputkey == 0:
                                        os.system('cls')
                                        print("대여 수공구 현황 확인 및 관리 종료")
                                        sleep(0.3)
                                        del_ref()
                                        break
                                    elif 0 < inputkey <= len(TOTAL_rental_list):
                                        os.system('cls')
                                        print(inputkey,".",WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=3).value,"-",TOTAL_rental_name)
                                        print("를 반납처리하겠습니까? 확인은 1, 취소는 1을 제외한 아무키나 입력하십시오")
                                        inputOX = input("")
                                        if inputOX == '1':
                                            os.system('cls')
                                            TL_rental_uid = WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=5).value
                                            TL_rental_row = findUID(WS_H,WS_H_COUNTER,1,TL_rental_uid)
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=TL_rental_uid)
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=WS_H.cell(row=TL_rental_row,column=2).value)
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=TL_rental_row,column=3).value)
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=TL_rental_row,column=5).value)
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="강제반납")
                                            WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=3).value)
                                            WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=4,value=0)
                                            WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=5,value="")
                                            WS_LOG_COUNTER += 1                                            
                                            wb.save(excel_name)
                                            wb2.save(excel_log)
                                            print(WS_TL.cell(row=TOTAL_rental_list[inputkey-1],column=3).value,"를 반납했습니다")
                                            sleep(0.3)
                                            del_ref()
                                        else:
                                            os.system('cls')                                            
                                            print("취소되었습니다.")
                                            sleep(0.3)
                                            del_ref()
                                    else:
                                        os.system('cls')                                        
                                        print("잘못된 입력입니다.")
                                        sleep(0.3)
                                        del_ref()
                            TOTAL_rental_list = []



                # 3-3. 관리자 모드 - 신규 등록 명단 열람
                elif admin == '3':
                    os.system('cls')
                    print("신규 등록 인원 확인\n")
                    if NewH_counter == 0:
                        print("신규 등록이 없습니다")
                        sleep(0.8)
                        del_ref()
                    else:    
                        print("신규 등록 인원: ",NewH_counter,"명")
                        Range_New = range(NewH_counter,0,-1)
                        print("신규 등록 명단:\n")
                        for i in Range_New:
                            if i>0:
                                CALL_H(WS_H_COUNTER-i)
                        print("확인이 끝났으면 아무키나 입력해주십시오")
                        NewH_counter = 0
                        input("")
                    

                # 3-4. 관리자 모드 - 나가기
                elif admin == '4':
                    os.system('cls')
                    break
                
                else : 
                    print("잘못된 입력입니다")
                    del_ref()


    # Main - 1,2,3,4 이외의 경우에 대한 입력의 경우 
    else:
        print("잘못된 입력입니다")
        del_ref()