import tkinter as tk                
from tkinter import font  as tkfont 
import tkinter.ttk 
import tkinter.messagebox as msgbox
import openpyxl

from time import time
from time import sleep
from time import strftime
from time import localtime
# 엑셀 파일
excel_name = 'TAGDATA.xlsx'
excel_log = 'TAGLOG.xlsx'
wb = openpyxl.load_workbook(filename=excel_name)         # filename= 에 엑셀 data 파일, wb: 엑셀 파일 자체
wb2 = openpyxl.load_workbook(filename=excel_log)
# 엑셀의 WorkSheet
WS_H = wb['사람']                                         
WS_TL = wb['공구']
WS_LOG = wb2['기록']
WS_ACCESSLIST = wb2['출입명부']
WS_3D = wb2['3D프린터이용대장']
WS_3DLIST = wb['3D프린터명단']
WS_PW = wb['관리자메뉴비밀번호']
# DB 업데이트용 마지막 줄 COUNTER 등록
WS_H_COUNTER= WS_H.max_row + 1                           
WS_TL_COUNTER = WS_TL.max_row + 1
WS_LOG_COUNTER = WS_LOG.max_row + 1
WS_ACCESSLIST_COUNTER = WS_ACCESSLIST.max_row + 1
WS_3D_COUNTER = WS_3D.max_row + 1
WS_3DLIST_COUNTER = WS_3DLIST.max_row + 1
# 관리자 비밀번호
password = WS_PW.cell(row=1,column=2).value  
# GUI 클릭 및 엔터 카운트 하는 용도 -> 버튼 클릭 한번으로 하나의 실행만 하게끔
NewH_counter = 0
Entercheck = 0
Btncheck = 0
def initEntercheck():
    global Entercheck
    Entercheck = 0

def countEntercheck():
    global Entercheck
    Entercheck +=1

def initBtncheck():
    global Btncheck
    Btncheck = 0

def countBtncheck():
    global Btncheck
    Btncheck +=1


def findUID(WS,WS_COUNTER,col,UID):         # UID 검색 함수(WS: WS 종류, WS_COUNTER: WS COUNTER 종류, col: 행, UID: 검색 UID)
    Rang = range(2, WS_COUNTER)             # 함수 결과는 UID가 있으면 UID가 몇번째 열에 있는지 나옴
    location = 2
    if WS_COUNTER == 2:
        return -1                           # UID가 DB에 없다면 -1 출력
    else:
        for i in Rang:
            if i < WS_COUNTER-1:
                if UID == WS.cell(row=location,column=col).value:
                    return location
                else:
                    location = location+1
                    continue
            elif i == WS_COUNTER-1:
                if UID == WS.cell(row=location,column=col).value:
                    return location
                else:
                    return -1

def CALL_H(WS_H_NUMBER):                    # 사용자 DB에서 WS_H_NUMBER 행의 정보를 불러오는 함수
    H_UID = str(WS_H.cell(row=WS_H_NUMBER, column=1).value)
    H_NAME = str(WS_H.cell(row=WS_H_NUMBER, column=2).value)
    H_stdID = str(WS_H.cell(row=WS_H_NUMBER, column=3).value)
    H_Phone = str(WS_H.cell(row=WS_H_NUMBER, column=4).value)
    H_depart = str(WS_H.cell(row=WS_H_NUMBER, column=5).value)
    H_info = H_UID+"    "+H_NAME+"    "+H_depart+"    "+H_stdID+"    "+H_Phone
    return H_info


def CALL_TIME():                            # 로그 시간 기록
    tm = localtime(time())
    return strftime('%Y-%m-%d %I:%M:%S %p', tm)

def CALL_TIME2():                           # 3D 프린터용 시간 기록 
    tm = localtime(time())
    return strftime('%Y-%m-%d',tm)

# MainStream : GUI frame을 제작하는 메인 툴
class MainStream(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("TAG")
        self.geometry("1920x1080+0+0")
        self.title_font = tkfont.Font(family='맑은 고딕', size=20, weight="bold")
        self.home_btn = tk.PhotoImage(file = "return.png")
        self._frame = None
        self.switch_frame(Home)
    # frame 간 화면 전환 시 destroy 후 rebuild하는 형식으로 switch
    def switch_frame(self, frame_class):
        new_frame = frame_class(self)
        if self._frame is not None:
            self._frame.destroy()
        self._frame = new_frame
        self._frame.pack()
    

# Main - Home 화면 
class Home(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경화면
        self.bg_img = tk.PhotoImage(file="Home.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Settings_Check로 가는 버튼 생성
        self.setbtn = tk.PhotoImage(file="setbtn.png")
        set_btn = tk.Button(self, width= 100, height= 100, bd=0,image=self.setbtn, command=lambda: master.switch_frame(Settings_Check) )
        set_btn.place(x=30,y=30)
        # 한/영 전환
        self.engbtn = tk.PhotoImage(file="eng.png")
        eng_btn = tk.Button(self, width= 100, height= 100, bd=0,image=self.engbtn, command=lambda: master.switch_frame(Home_EN))  
        eng_btn.place(x=160,y=30)
        # 4가지 기능을 각각 담당하는 버튼들 생성
        # 1. 출입 명부 작성 / 2. 3D프린터 이용대장 작성 / 3. 수공구 대여 / 4. 신규 등록
        self.selectbtn1 = tk.PhotoImage(file="main1_btn.png")
        self.selectbtn2 = tk.PhotoImage(file="main2_btn.png")
        self.selectbtn3 = tk.PhotoImage(file="main3_btn.png")
        self.selectbtn4 = tk.PhotoImage(file="main4_btn.png")
        button1 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn1 ,command=lambda: master.switch_frame(PageOne))
        button2 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn2 ,  command=lambda: master.switch_frame(PageTwo))
        button3 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn3 ,  command=lambda: master.switch_frame(PageThree))
        button4 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn4 ,  command=lambda: master.switch_frame(PageFour))

        button1.place(x=1050,y=60)
        button2.place(x=1050,y=290)
        button3.place(x=1050,y=520)
        button4.place(x=1050,y=750)

# Main - 1. 출입 명부 작성
class PageOne(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="1출입명부.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)

        label2=tk.Label(self,font=master.title_font,fg='white',bg='#3B3838')

        # 방문 목적 1,2 에 맞는지 확인하는 경고창 생성 후 확인 시 저장
        def askq(text,H_name,H_location,WS_ACCESSLIST_COUNTER_FN):
            msg = msgbox.askquestion(text+"목적",H_name+"님\n" +text+ "목적으로 방문하시겠습니까?")
            if msg == 'yes':
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=1,value=WS_ACCESSLIST_COUNTER_FN-1)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=2,value=CALL_TIME())
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=3,value=H_name)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=4,value=WS_H.cell(row=H_location, column=5).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=5,value=WS_H.cell(row=H_location, column=4).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=6,value=text)
                wb2.save(excel_log)
                global WS_ACCESSLIST_COUNTER
                WS_ACCESSLIST_COUNTER +=1
                master.switch_frame(Home)
        # 방문 목적 기타인 경우에 대한 askq 함수의 맞춤 ver.
        def askq2(text,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,etcwindow,master):
            msg = msgbox.askquestion(text+"목적",H_name+"님\n" +text+ " 목적으로 방문하시겠습니까?")
            if msg == 'yes':
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=1,value=WS_ACCESSLIST_COUNTER_FN-1)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=2,value=CALL_TIME())
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=3,value=H_name)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=4,value=WS_H.cell(row=H_location, column=5).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=5,value=WS_H.cell(row=H_location, column=4).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=6,value=text)
                wb2.save(excel_log)
                global WS_ACCESSLIST_COUNTER
                WS_ACCESSLIST_COUNTER += 1
                master.switch_frame(Home)
                etcwindow.withdraw()

        # 기타 목적 선택 시 추가 창 생성 및 기타 목적 저장
        self.empty_img = tk.PhotoImage(file="blank.png")

        def etc(self,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,master):
            def etcentry(n):
                etcpurp = etcent.get()
                askq2(etcpurp,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,etcwindow,master)
            etcwindow = tk.Toplevel(self)
            etcwindow.geometry("1920x1080+0+0")
            sky = tk.Label(etcwindow,image=self.empty_img)
            sky.pack(fill="both")
            etclbl = tk.Label(etcwindow, text="기타 목적을 입력해주십시오\n\n입력 완료 시 엔터를 눌러주십시오",font=master.title_font,fg='white',bg='#3B3838')
            etcent = tk.Entry(etcwindow, width=30)
            etclbl.place(x=750,y=200)
            etcent.place(x=850,y=340)
            etcent.bind("<Return>",etcentry)

        self.btn1_img = tk.PhotoImage(file="1출입_테이블사용.png")
        self.btn2_img = tk.PhotoImage(file="1출입_유저동행.png")
        self.btn3_img = tk.PhotoImage(file="1출입_기타.png")
        
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("경고!","등록되지 않은 사용자입니다. 신규등록을 해주십시오")
                else:
                    countEntercheck()
                    H_name = WS_H.cell(row=H_location, column=2).value
                    label2.config(text=H_name+"님 환영합니다.\n방문목적을 선택해주십시오\n")
                    label2.place(x=790,y=400)
                    # 1. 테이블 사용 / 2. 유저 동행 / 3. 기타
                    btn1 = tk.Button(self, width= 800, height=100 , image=self.btn1_img, bd=0,command=lambda: askq("테이블 사용 ",H_name,H_location,WS_ACCESSLIST_COUNTER))
                    btn2 = tk.Button(self, width= 800, height=100 , image=self.btn2_img, bd=0, command=lambda: askq("유저 동행 ",H_name,H_location,WS_ACCESSLIST_COUNTER))
                    btn3 = tk.Button(self, width= 800, height=100 , image=self.btn3_img, bd=0, command=lambda: etc(self,H_name,H_location,WS_ACCESSLIST_COUNTER,master)) 
                    
                    btn1.place(x=560,y=530)
                    btn2.place(x=560,y=670)
                    btn3.place(x=560,y=810)
        
        entry = tk.Entry(self,width=30,show='*')
        # 엔터키 (<Return>) 을 입력 시 result 함수 실행
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# Main - 2. 3D 프린터 이용대장 작성       
class PageTwo(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        initBtncheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="23d프린터.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)

        self.btn = tk.PhotoImage(file="완료.png")
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("경고!","등록되지 않은 사용자입니다. 신규등록을 해주십시오")
                else:
                    countEntercheck()
                    H_name = WS_H.cell(row=H_location, column=2).value
                    # 현재 KU Makerspace에 있는 3D 프린터 나열
                    printer_list = []
                    LIST3D_RANGE = range(WS_3DLIST_COUNTER-2)
                    for i in LIST3D_RANGE:
                        printer_list.append(WS_3DLIST.cell(row=i+2, column=1).value)
                    
                    label2 = tk.Label(self,text=H_name+" 님 환영합니다\n3D 프린터명을 입력해주십시오",font=master.title_font,fg='white',bg='#3B3838')
                    choicebox = tkinter.ttk.Combobox(self, values= printer_list,state='readonly',font=master.title_font)
                    choicebox.set("3D 프린터 선택")
                    
                    label2.place(x=780,y=400)
                    choicebox.place(x=800,y=530)

                    # 3D 프린터 사용 시간 10:00-18:00 30분 단위
                    start = ['10:00','10:30','11:00','11:30','12:00','12:30','13:00','13:30',
                            '14:00','14:30','15:00','15:30','16:00','16:30','17:00','17:30','18:00']
                    starttime = tkinter.ttk.Combobox(self,height=5, values=start[:16],state='readonly',font=master.title_font)
                    starttime.place(x=800,y=590)
                    # 시작 시간 선택
                    def starttimecheck():
                        if Btncheck == 0:
                            global startcheck, endcheck, printerchoice
                            printerchoice = choicebox.get()
                            startcheck = starttime.get()
                            if startcheck != "":
                                for i in range(0,16):
                                    if startcheck == start[i]:
                                        endtime = tkinter.ttk.Combobox(self,height=5,values=start[i+1:],state='readonly',font=master.title_font)
                                        endcheck = endtime.get()
                                        endtime.place(x=800,y=740)
                                        finbtn = tk.Button(self,width=160,height=90,bd=0,image=self.btn,font=master.title_font,command=lambda:endtimecheck(endtime))
                                        finbtn.place(x=880,y=800)
                                        countBtncheck()
                    # 시작 시간 선택을 기반으로 그 시간 +30분 이후부터 종료 시간 선택 가능
                    def endtimecheck(endtime):
                        endcheck = endtime.get()
                        printerchoice = choicebox.get()
                        if printerchoice != "3D 프린터 선택":
                            tottime = str(startcheck+" ~ "+endcheck)
                            confirm = msgbox.askquestion("3D 프린터 사용 등록","날짜: "+str(CALL_TIME2())+"\n이용자명: "+H_name+"\n프린터명: "+printerchoice+"\n사용시간: "+tottime)
                            if confirm == 'yes':
                                global WS_3D_COUNTER
                                WS_3D.cell(row=WS_3D_COUNTER,column=1,value=CALL_TIME2())
                                WS_3D.cell(row=WS_3D_COUNTER,column=2,value=WS_H.cell(row=H_location, column=5).value)
                                WS_3D.cell(row=WS_3D_COUNTER,column=3,value=WS_H.cell(row=H_location, column=3).value)
                                WS_3D.cell(row=WS_3D_COUNTER,column=4,value=H_name)
                                WS_3D.cell(row=WS_3D_COUNTER,column=5,value=printerchoice)
                                WS_3D.cell(row=WS_3D_COUNTER,column=6,value=tottime)
                                wb2.save(excel_log)
                                WS_3D_COUNTER +=1
                                master.switch_frame(Home)

                    startbtn = tk.Button(self,text="시작 시간 선택",fg='black',bg='#FFC000',command=starttimecheck,font=master.title_font)
                    startbtn.place(x=860,y=650)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# Main - 3. 수공구 대여/반납/대여확인
class PageThree(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="3수공구.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)

        label2=tk.Label(self)
        # 수공구 대여 시 
        self.brw_win = tk.PhotoImage(file='3수공구대여.png')
        def toolBrw(listbox,TL_rental_list,H_name,H_location):
        # listbox, TL_rental_list: 대여 현황 목록, H_name: 대여 사람 이름, H_location: 대여 사람 엑셀 열(row)
            Brw_win = tk.Toplevel(self)
            Brw_win.geometry("1920x1080+0+0")

            def toolBrw_confirm(n):
                tooluid = Brw_ent.get()
                TL_location = findUID(WS_TL,WS_TL_COUNTER,1,tooluid)
                if TL_location == -1:
                    Brw_ent.delete(0,'end')
                    msgbox.showwarning("경고!","등록되지 않은 수공구입니다!")
                    Brw_win.lift()
                elif WS_TL.cell(row=TL_location,column=4).value == 1:
                    Brw_ent.delete(0,'end')
                    msgbox.showwarning("경고!","이미 대여된 수공구입니다!")
                    Brw_win.lift()
                else:
                    tool_name = WS_TL.cell(row=TL_location,column=3).value
                    msg = msgbox.askquestion("대여 확인",tool_name+"를 대여하시겠습니까?",default='no')
                    if msg == 'yes':
                        # DATA 대여 기록
                        WS_TL.cell(row=TL_location,column=4,value=1)
                        WS_TL.cell(row=TL_location,column=5,value=entry.get())
                        # LOG 대여 기록
                        global WS_LOG_COUNTER
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=entry.get())
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=H_name)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="대여")
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=tool_name)
                        wb.save(excel_name)
                        wb2.save(excel_log)
                        WS_LOG_COUNTER += 1
                        TL_rental_list.append(tool_name)
                        listbox.insert(len(TL_rental_list),tool_name)
                        Brw_win.withdraw()
                        
            Brw_lbl = tk.Label(Brw_win, image=self.brw_win)
            Brw_ent = tk.Entry(Brw_win, width=30,show='*')
            Brw_lbl.pack(fill='both')
            Brw_ent.place(x=850,y=340)
            Brw_ent.bind("<Return>",toolBrw_confirm)
            
        # 수공구 반납 시
        self.rtn_win = tk.PhotoImage(file='3수공구반납.png')
        def toolRtn(listbox,TL_rental_list,Rtn_uid,Rtn_name,H_location):    
            # listbox, TL_rental_list: 대여 현황 목록, Rtn_uid: 반납 사람 UID, Rtn_name: 반납 사람 이름
            if listbox.size() == 0:
                msgbox.showwarning("경고!","반납할 수공구가 없습니다!")
            else:
            
                Rtn_win = tk.Toplevel(self)
                Rtn_win.geometry("1920x1080+0+0")
                Rtn_win_lbl = tk.Label(Rtn_win,image=self.rtn_win)
                Rtn_win_lbl.pack(fill='both')

                Rtn_win_ent = tk.Entry(Rtn_win,width=30,show='*')

                def toolRtn_confirm(n):
                    
                    tool_uid = Rtn_win_ent.get()
                    TL_location = findUID(WS_TL,WS_TL_COUNTER,1,tool_uid)
                    if TL_location == -1:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("경고!","등록되지 않은 수공구입니다!")
                        Rtn_win.lift()
                        
                    elif WS_TL.cell(row=TL_location,column=4).value == 0:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("경고!","이미 반납된 수공구입니다!")
                        Rtn_win.lift()
                        
                    elif WS_TL.cell(row=TL_location,column=5).value != Rtn_uid:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("경고!","사용자가 대여한 공구가 아닙니다!")
                        Rtn_win.lift()
                        
                    else:
                        index =0
                        tool_Rtn_name = WS_TL.cell(row=TL_location,column=3).value
                        for i in range(0,len(TL_rental_list)):
                            if tool_Rtn_name == TL_rental_list[i]:
                                index = i
                        msg = msgbox.askquestion("수공구 반납 확인",tool_Rtn_name+"를 반납하시겠습니까?",default='no')
                        if msg == 'yes':
                            # DATA 공구 반납 기록
                            WS_TL.cell(row=TL_location,column=4,value=0)
                            WS_TL.cell(row=TL_location,column=5,value="")
                            # LOG 반납 기록
                            global WS_LOG_COUNTER
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=Rtn_uid)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=Rtn_name)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="반납")
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=tool_Rtn_name)
                            wb.save(excel_name)
                            wb2.save(excel_log)
                            WS_LOG_COUNTER += 1
                            listbox.delete(index)
                            TL_rental_list.remove(tool_Rtn_name)
                            Rtn_win.withdraw()
                
                Rtn_win_ent.bind("<Return>",toolRtn_confirm)
                Rtn_win_ent.place(x=850,y=340)

        self.Brw_btn = tk.PhotoImage(file='3대여버튼.png')
        self.Rtn_btn = tk.PhotoImage(file='3반납버튼.png')
        def result(n):
            H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
            if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                entry.delete(0,'end')
                msgbox.showwarning("경고!","등록되지 않은 사용자입니다. 신규등록을 해주십시오")
            else:
                if Entercheck == 0:
                    H_name = WS_H.cell(row=H_location, column=2).value
                    label2.config(text=H_name+"님 환영합니다.\n현재 대여 현황은 다음과 같습니다\n",font=master.title_font,fg='white',bg='#3B3838')
                    countEntercheck()
                    label2.place(x=760,y=400)
                    # 대여 현황 띄우기
                    listbox = tk.Listbox(self,width=25,height=10,font=master.title_font)

                    TL_rental_list = []
                    TL_rental_range = range(2,WS_TL_COUNTER)
                    for i in TL_rental_range:
                        if entry.get() == WS_TL.cell(row=i,column=5).value:     # TAG 한 학생증 UID authorization
                            brw_item = WS_TL.cell(row=i,column=3).value
                            TL_rental_list.append(brw_item)
                    # 스크롤바 생성
                    for i in range(0,len(TL_rental_list)):
                        listbox.insert(i,TL_rental_list[i])
                    listbox.place(x=500,y=520)
                    
                    btn_brw = tk.Button(self,width=400,height=150,bd=0,image=self.Brw_btn,command=lambda:toolBrw(listbox,TL_rental_list,H_name,H_location))
                    btn_rtn = tk.Button(self,width=400,height=150,bd=0,image=self.Rtn_btn,command=lambda:toolRtn(listbox,TL_rental_list,entry.get(),H_name,H_location))

                    btn_brw.place(x=1010,y=535)
                    btn_rtn.place(x=1010,y=735)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)
        

# Main - 4. 신규등록
class PageFour(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initBtncheck()
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="4신규.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)

        self.comp_btn = tk.PhotoImage(file='완료.png')
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location != -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("경고!","이미 등록된 사용자입니다")
                elif findUID(WS_TL,WS_TL_COUNTER,1,entry.get()) != -1:
                    entry.delete(0,'end')
                    msgbox.showwarning("경고!","이미 등록된 RFID 입니다.")
                else:
                    countEntercheck()
                    lbl_name = tk.Label(self,text="이름",font=master.title_font,fg='white',bg='#3B3838')
                    ent_name = tk.Entry(self,font=master.title_font)
                    lbl_depart = tk.Label(self,text="학과",font=master.title_font,fg='white',bg='#3B3838')
                    ent_depart = tk.Entry(self,font=master.title_font)
                    lbl_stdnum = tk.Label(self,text="학번",font=master.title_font,fg='white',bg='#3B3838')
                    ent_stdnum = tk.Entry(self,font=master.title_font)
                    lbl_phone = tk.Label(self,text="전화번호 (-는 생략)",font=master.title_font,fg='white',bg='#3B3838')
                    ent_phone = tk.Entry(self,font=master.title_font)
                    lbl_name.place(x=930,y=400)
                    ent_name.place(x=800,y=450)
                    lbl_depart.place(x=930,y=500)
                    ent_depart.place(x=800,y=550)
                    lbl_stdnum.place(x=930,y=600)
                    ent_stdnum.place(x=800,y=650)
                    lbl_phone.place(x=850,y=700)
                    ent_phone.place(x=800,y=750)
                    
                    def newregister():
                        uid = entry.get()
                        name = ent_name.get()
                        depart = ent_depart.get()
                        stdnum = ent_stdnum.get()
                        phone = ent_phone.get()
                        if len(name)!=0 and len(depart)!=0 and len(stdnum)!=0 and len(phone) != 0:
                            confirm_msg = msgbox.askquestion("신규 등록 확인","이름 : "+ name+"\n학과 : "+depart+"\n학번 : "+stdnum+"\n전화번호 : "+phone)
                        
                            if confirm_msg == 'yes':
                                global WS_H_COUNTER
                                global NewH_counter
                                WS_H.cell(row=WS_H_COUNTER,column=1,value=uid)
                                WS_H.cell(row=WS_H_COUNTER,column=2,value=name)
                                WS_H.cell(row=WS_H_COUNTER,column=3,value=stdnum)
                                WS_H.cell(row=WS_H_COUNTER,column=4,value=phone)
                                WS_H.cell(row=WS_H_COUNTER,column=5,value=depart)
                                wb.save(excel_name)
                                WS_H_COUNTER += 1
                                NewH_counter += 1
                                
                                master.switch_frame(Home)

                    newreg = tk.Button(self,width=160,height=90,image=self.comp_btn,command=lambda:newregister())
                    newreg.place(x=880,y=850)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# EN ver. of TAG
# Main - Home Window 
class Home_EN(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경화면
        self.bg_img = tk.PhotoImage(file="Home.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Settings_Check로 가는 버튼 생성
        self.setbtn = tk.PhotoImage(file="setbtn.png")
        set_btn = tk.Button(self, width= 100, height= 100, bd=0,image=self.setbtn, command=lambda: master.switch_frame(Settings_Check) )
        set_btn.place(x=30,y=30)
        # 한글 전환
        self.engbtn = tk.PhotoImage(file="한글.png")
        eng_btn = tk.Button(self, width= 100, height= 100, bd=0,image=self.engbtn, command=lambda: master.switch_frame(Home))  
        eng_btn.place(x=160,y=30)
        # 4가지 기능을 각각 담당하는 버튼들 생성
        # 1. 출입 명부 작성 / 2. 3D프린터 이용대장 작성 / 3. 수공구 대여 / 4. 신규 등록
        self.selectbtn1 = tk.PhotoImage(file="main1GuestBook.png")
        self.selectbtn2 = tk.PhotoImage(file="main23dprinter.png")
        self.selectbtn3 = tk.PhotoImage(file="main3tool.png")
        self.selectbtn4 = tk.PhotoImage(file="main4newreg.png")
        button1 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn1 ,command=lambda: master.switch_frame(PageOne_EN))
        button2 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn2 ,  command=lambda: master.switch_frame(PageTwo_EN))
        button3 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn3 ,  command=lambda: master.switch_frame(PageThree_EN))
        button4 = tk.Button(self, width= 800, height=200, bd=1, image=self.selectbtn4 ,  command=lambda: master.switch_frame(PageFour_EN))

        button1.place(x=1050,y=60)
        button2.place(x=1050,y=290)
        button3.place(x=1050,y=520)
        button4.place(x=1050,y=750)


# Main - 1. Guest Book
class PageOne_EN(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="1Guestbook.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home_EN))
        home_button.place(x=30,y=30)

        label2=tk.Label(self,font=master.title_font,fg='white',bg='#3B3838')

        # 방문 목적 1,2 에 맞는지 확인하는 경고창 생성 후 확인 시 저장
        def askq(text,H_name,H_location,WS_ACCESSLIST_COUNTER_FN):
            msg = msgbox.askquestion("For "+text,H_name+"\n" +"Please confirm your purpose of visit\nPurpose: "+text)
            if msg == 'yes':
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=1,value=WS_ACCESSLIST_COUNTER_FN-1)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=2,value=CALL_TIME())
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=3,value=H_name)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=4,value=WS_H.cell(row=H_location, column=5).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=5,value=WS_H.cell(row=H_location, column=4).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=6,value=text)
                wb2.save(excel_log)
                global WS_ACCESSLIST_COUNTER
                WS_ACCESSLIST_COUNTER +=1
                master.switch_frame(Home_EN)
        # 방문 목적 기타인 경우에 대한 askq 함수의 맞춤 ver.
        def askq2(text,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,etcwindow,master):
            msg = msgbox.askquestion("For "+text,H_name+"\n" +"Please confirm your purpose of visit\nPurpose: "+text)
            if msg == 'yes':
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=1,value=WS_ACCESSLIST_COUNTER_FN-1)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=2,value=CALL_TIME())
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=3,value=H_name)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=4,value=WS_H.cell(row=H_location, column=5).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=5,value=WS_H.cell(row=H_location, column=4).value)
                WS_ACCESSLIST.cell(row=WS_ACCESSLIST_COUNTER_FN,column=6,value=text)
                wb2.save(excel_log)
                global WS_ACCESSLIST_COUNTER
                WS_ACCESSLIST_COUNTER += 1
                master.switch_frame(Home_EN)
                etcwindow.withdraw()

        # 기타 목적 선택 시 추가 창 생성 및 기타 목적 저장
        self.empty_img = tk.PhotoImage(file="blank.png")

        def etc(self,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,master):
            def etcentry(n):
                etcpurp = etcent.get()
                askq2(etcpurp,H_name,H_location,WS_ACCESSLIST_COUNTER_FN,etcwindow,master)
            etcwindow = tk.Toplevel(self)
            etcwindow.geometry("1920x1080+0+0")
            sky = tk.Label(etcwindow,image=self.empty_img)
            sky.pack(fill="both")
            etclbl = tk.Label(etcwindow, text="Please type in your personal purpose\n\nPress Enter when finished",font=master.title_font,fg='white',bg='#3B3838')
            etcent = tk.Entry(etcwindow, width=30)
            etclbl.place(x=720,y=200)
            etcent.place(x=850,y=340)
            etcent.bind("<Return>",etcentry)

        self.btn1_img = tk.PhotoImage(file="1Gb_Table.png")
        self.btn2_img = tk.PhotoImage(file="1Gb_Accomp.png")
        self.btn3_img = tk.PhotoImage(file="1Gb_Etc.png")
        
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("Warning!","User has not been registered. Please proceed to new registration")
                else:
                    countEntercheck()
                    H_name = WS_H.cell(row=H_location, column=2).value
                    label2.config(text="Welcome! "+H_name+"\nPlease select the purpose of your visit\n")
                    label2.place(x=720,y=400)
                    # 1. 테이블 사용 / 2. 유저 동행 / 3. 기타
                    btn1 = tk.Button(self, width= 800, height=100 , image=self.btn1_img, bd=0,command=lambda: askq("Table Usage ",H_name,H_location,WS_ACCESSLIST_COUNTER))
                    btn2 = tk.Button(self, width= 800, height=100 , image=self.btn2_img, bd=0, command=lambda: askq("Accompany ",H_name,H_location,WS_ACCESSLIST_COUNTER))
                    btn3 = tk.Button(self, width= 800, height=100 , image=self.btn3_img, bd=0, command=lambda: etc(self,H_name,H_location,WS_ACCESSLIST_COUNTER,master)) 
                    
                    btn1.place(x=560,y=530)
                    btn2.place(x=560,y=670)
                    btn3.place(x=560,y=810)
        
        entry = tk.Entry(self,width=30,show='*')
        # 엔터키 (<Return>) 을 입력 시 result 함수 실행
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# Main - 2. 3D Printer Usage Record       
class PageTwo_EN(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        initBtncheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="23dprinter.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home_EN))
        home_button.place(x=30,y=30)

        self.btn = tk.PhotoImage(file="OK.png")
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("Warning!","User has not been registered. Please proceed to new registration")
                else:
                    countEntercheck()
                    H_name = WS_H.cell(row=H_location, column=2).value
                    # 현재 KU Makerspace에 있는 3D 프린터 나열
                    printer_list = []
                    LIST3D_RANGE = range(WS_3DLIST_COUNTER-2)
                    for i in LIST3D_RANGE:
                        printer_list.append(WS_3DLIST.cell(row=i+2, column=1).value)
                    
                    label2 = tk.Label(self,text="Welcome! "+H_name+"\nPlease choose your reserved 3D Printer",font=master.title_font,fg='white',bg='#3B3838')
                    choicebox = tkinter.ttk.Combobox(self, values= printer_list,state='readonly',font=master.title_font)
                    choicebox.set("Choose 3D Printer")
                    
                    label2.place(x=700,y=400)
                    choicebox.place(x=800,y=530)

                    # 3D 프린터 사용 시간 10:00-18:00 30분 단위
                    start = ['10:00','10:30','11:00','11:30','12:00','12:30','13:00','13:30',
                            '14:00','14:30','15:00','15:30','16:00','16:30','17:00','17:30','18:00']
                    starttime = tkinter.ttk.Combobox(self,height=5, values=start[:16],state='readonly',font=master.title_font)
                    starttime.place(x=800,y=590)
                    # 시작 시간 선택
                    def starttimecheck():
                        if Btncheck == 0:
                            global startcheck, endcheck, printerchoice
                            printerchoice = choicebox.get()
                            startcheck = starttime.get()
                            if startcheck != "":
                                for i in range(0,16):
                                    if startcheck == start[i]:
                                        endtime = tkinter.ttk.Combobox(self,height=5,values=start[i+1:],state='readonly',font=master.title_font)
                                        endcheck = endtime.get()
                                        endtime.place(x=800,y=740)
                                        finbtn = tk.Button(self,width=160,height=90,bd=0,image=self.btn,font=master.title_font,command=lambda:endtimecheck(endtime))
                                        finbtn.place(x=880,y=800)
                                        countBtncheck()
                    # 시작 시간 선택을 기반으로 그 시간 +30분 이후부터 종료 시간 선택 가능
                    def endtimecheck(endtime):
                        endcheck = endtime.get()
                        printerchoice = choicebox.get()
                        if printerchoice != "Choose 3D Printer":
                            tottime = str(startcheck+" ~ "+endcheck)
                            confirm = msgbox.askquestion("Register 3D Printer choice",
                                                        "Date: "+str(CALL_TIME2())+"\nName: "+H_name+"\n3D Printer: "+printerchoice+"\nTime: "+tottime)
                            if confirm == 'yes':
                                global WS_3D_COUNTER
                                WS_3D.cell(row=WS_3D_COUNTER,column=1,value=CALL_TIME2())
                                WS_3D.cell(row=WS_3D_COUNTER,column=2,value=WS_H.cell(row=H_location, column=5).value)
                                WS_3D.cell(row=WS_3D_COUNTER,column=3,value=WS_H.cell(row=H_location, column=3).value)
                                WS_3D.cell(row=WS_3D_COUNTER,column=4,value=H_name)
                                WS_3D.cell(row=WS_3D_COUNTER,column=5,value=printerchoice)
                                WS_3D.cell(row=WS_3D_COUNTER,column=6,value=tottime)
                                wb2.save(excel_log)
                                WS_3D_COUNTER +=1
                                master.switch_frame(Home_EN)

                    startbtn = tk.Button(self,text="Select start time",fg='black',bg='#FFC000',command=starttimecheck,font=master.title_font)
                    startbtn.place(x=850,y=650)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# Main - 3. Handtool Borrow/Return
class PageThree_EN(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="3tool.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home_EN))
        home_button.place(x=30,y=30)

        label2=tk.Label(self)
        # 수공구 대여 시 
        self.brw_win = tk.PhotoImage(file='3brw.png')
        def toolBrw(listbox,TL_rental_list,H_name,H_location):
        # listbox, TL_rental_list: 대여 현황 목록, H_name: 대여 사람 이름, H_location: 대여 사람 엑셀 열(row)
            Brw_win = tk.Toplevel(self)
            Brw_win.geometry("1920x1080+0+0")

            def toolBrw_confirm(n):
                tooluid = Brw_ent.get()
                TL_location = findUID(WS_TL,WS_TL_COUNTER,1,tooluid)
                if TL_location == -1:
                    Brw_ent.delete(0,'end')
                    msgbox.showwarning("Warning!","Tool not registered!")
                    Brw_win.lift()
                elif WS_TL.cell(row=TL_location,column=4).value == 1:
                    Brw_ent.delete(0,'end')
                    msgbox.showwarning("Warning!","This tool has been already borrowed!")
                    Brw_win.lift()
                else:
                    tool_name = WS_TL.cell(row=TL_location,column=3).value
                    msg = msgbox.askquestion("Confirm tool borrowing","Will you borrow "+tool_name+"?",default='no')
                    if msg == 'yes':
                        # DATA 대여 기록
                        WS_TL.cell(row=TL_location,column=4,value=1)
                        WS_TL.cell(row=TL_location,column=5,value=entry.get())
                        # LOG 대여 기록
                        global WS_LOG_COUNTER
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=entry.get())
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=H_name)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="대여")
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=tool_name)
                        wb.save(excel_name)
                        wb2.save(excel_log)
                        WS_LOG_COUNTER += 1
                        TL_rental_list.append(tool_name)
                        listbox.insert(len(TL_rental_list),tool_name)
                        Brw_win.withdraw()
                        
            Brw_lbl = tk.Label(Brw_win, image=self.brw_win)
            Brw_ent = tk.Entry(Brw_win, width=30,show='*')
            Brw_lbl.pack(fill='both')
            Brw_ent.place(x=850,y=340)
            Brw_ent.bind("<Return>",toolBrw_confirm)
            
        # 수공구 반납 시
        self.rtn_win = tk.PhotoImage(file='3rtn.png')
        def toolRtn(listbox,TL_rental_list,Rtn_uid,Rtn_name,H_location):    
            # listbox, TL_rental_list: 대여 현황 목록, Rtn_uid: 반납 사람 UID, Rtn_name: 반납 사람 이름
            if listbox.size() == 0:
                msgbox.showwarning("Warning!","You have no tools to return!")
            else:
                Rtn_win = tk.Toplevel(self)
                Rtn_win.geometry("1920x1080+0+0")
                Rtn_win_lbl = tk.Label(Rtn_win,image=self.rtn_win)
                Rtn_win_lbl.pack(fill='both')

                Rtn_win_ent = tk.Entry(Rtn_win,width=30,show='*')

                def toolRtn_confirm(n):
                    
                    tool_uid = Rtn_win_ent.get()
                    TL_location = findUID(WS_TL,WS_TL_COUNTER,1,tool_uid)
                    if TL_location == -1:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("Warning!","Tool not registered!")
                        Rtn_win.lift()
                        
                    elif WS_TL.cell(row=TL_location,column=4).value == 0:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("Warning!","This tool has been already returned!")
                        Rtn_win.lift()
                        
                    elif WS_TL.cell(row=TL_location,column=5).value != Rtn_uid:
                        Rtn_win_ent.delete(0,'end')
                        msgbox.showwarning("Warning!","This tool has not been borrowed from you!")
                        Rtn_win.lift()
                        
                    else:
                        index =0
                        tool_Rtn_name = WS_TL.cell(row=TL_location,column=3).value
                        for i in range(0,len(TL_rental_list)):
                            if tool_Rtn_name == TL_rental_list[i]:
                                index = i
                        msg = msgbox.askquestion("Confirm tool returning","Will you return "+tool_Rtn_name+"?",default='no')
                        if msg == 'yes':
                            # DATA 공구 반납 기록
                            WS_TL.cell(row=TL_location,column=4,value=0)
                            WS_TL.cell(row=TL_location,column=5,value="")
                            # LOG 반납 기록
                            global WS_LOG_COUNTER
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=Rtn_uid)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=Rtn_name)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value)
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="반납")
                            WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=tool_Rtn_name)
                            wb.save(excel_name)
                            wb2.save(excel_log)
                            WS_LOG_COUNTER += 1
                            listbox.delete(index)
                            TL_rental_list.remove(tool_Rtn_name)
                            Rtn_win.withdraw()
                
                Rtn_win_ent.bind("<Return>",toolRtn_confirm)
                Rtn_win_ent.place(x=850,y=340)

        self.Brw_btn = tk.PhotoImage(file='3tool_brw.png')
        self.Rtn_btn = tk.PhotoImage(file='3tool_rtn.png')
        def result(n):
            H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
            if H_location == -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                entry.delete(0,'end')
                msgbox.showwarning("Warning!","User has not been registered. Please proceed to new registration")
            else:
                if Entercheck == 0:
                    H_name = WS_H.cell(row=H_location, column=2).value
                    label2.config(text="Welcome! "+H_name+"\nCurrent state of your rental list\n",font=master.title_font,fg='white',bg='#3B3838')
                    countEntercheck()
                    label2.place(x=760,y=400)
                    # 대여 현황 띄우기
                    listbox = tk.Listbox(self,width=25,height=10,font=master.title_font)

                    TL_rental_list = []
                    TL_rental_range = range(2,WS_TL_COUNTER)
                    for i in TL_rental_range:
                        if entry.get() == WS_TL.cell(row=i,column=5).value:     # TAG 한 학생증 UID authorization
                            brw_item = WS_TL.cell(row=i,column=3).value
                            TL_rental_list.append(brw_item)
                    # 스크롤바 생성
                    for i in range(0,len(TL_rental_list)):
                        listbox.insert(i,TL_rental_list[i])
                    listbox.place(x=500,y=520)
                    
                    btn_brw = tk.Button(self,width=400,height=150,bd=0,image=self.Brw_btn,command=lambda:toolBrw(listbox,TL_rental_list,H_name,H_location))
                    btn_rtn = tk.Button(self,width=400,height=150,bd=0,image=self.Rtn_btn,command=lambda:toolRtn(listbox,TL_rental_list,entry.get(),H_name,H_location))

                    btn_brw.place(x=1010,y=535)
                    btn_rtn.place(x=1010,y=735)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)
        

# Main - 4. New Registration
class PageFour_EN(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initBtncheck()
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="4newreg.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home_EN))
        home_button.place(x=30,y=30)

        self.comp_btn = tk.PhotoImage(file='OK.png')
        def result(n):
            if Entercheck == 0:
                H_location = findUID(WS_H,WS_H_COUNTER,1,entry.get())
                if H_location != -1:                # 잘못된 정보 입력시 자동 초기화 및 경고창 출력
                    entry.delete(0,'end')
                    msgbox.showwarning("Warning!","User has been already registered!")
                elif findUID(WS_TL,WS_TL_COUNTER,1,entry.get()) != -1:
                    entry.delete(0,'end')
                    msgbox.showwarning("Warning!","This RFID has been already registerd!")
                else:
                    countEntercheck()
                    lbl_name = tk.Label(self,text="Name",font=master.title_font,fg='white',bg='#3B3838')
                    ent_name = tk.Entry(self,font=master.title_font)
                    lbl_depart = tk.Label(self,text="Department",font=master.title_font,fg='white',bg='#3B3838')
                    ent_depart = tk.Entry(self,font=master.title_font)
                    lbl_stdnum = tk.Label(self,text="Student ID",font=master.title_font,fg='white',bg='#3B3838')
                    ent_stdnum = tk.Entry(self,font=master.title_font)
                    lbl_phone = tk.Label(self,text="Phone number (without -)",font=master.title_font,fg='white',bg='#3B3838')
                    ent_phone = tk.Entry(self,font=master.title_font)
                    lbl_name.place(x=920,y=400)
                    ent_name.place(x=800,y=450)
                    lbl_depart.place(x=880,y=500)
                    ent_depart.place(x=800,y=550)
                    lbl_stdnum.place(x=890,y=600)
                    ent_stdnum.place(x=800,y=650)
                    lbl_phone.place(x=790,y=700)
                    ent_phone.place(x=800,y=750)
                    
                    def newregister():
                        uid = entry.get()
                        name = ent_name.get()
                        depart = ent_depart.get()
                        stdnum = ent_stdnum.get()
                        phone = ent_phone.get()
                        if len(name)!=0 and len(depart)!=0 and len(stdnum)!=0 and len(phone) != 0:
                            confirm_msg = msgbox.askquestion("New Registration","Name : "+ name+"\nDepartment : "+depart+"\nStudent ID : "+stdnum+"\nPhone number : "+phone)
                            if Btncheck == 0:
                                if confirm_msg == 'yes':
                                    global WS_H_COUNTER
                                    global NewH_counter
                                    WS_H.cell(row=WS_H_COUNTER,column=1,value=uid)
                                    WS_H.cell(row=WS_H_COUNTER,column=2,value=name)
                                    WS_H.cell(row=WS_H_COUNTER,column=3,value=stdnum)
                                    WS_H.cell(row=WS_H_COUNTER,column=4,value=phone)
                                    WS_H.cell(row=WS_H_COUNTER,column=5,value=depart)
                                    wb.save(excel_name)
                                    WS_H_COUNTER += 1
                                    NewH_counter += 1
                                    countBtncheck()
                                    master.switch_frame(Home_EN)

                    newreg = tk.Button(self,width=160,height=90,image=self.comp_btn,command=lambda:newregister())
                    newreg.place(x=880,y=850)

        entry = tk.Entry(self,width=30,show='*')
        entry.bind("<Return>",result)
        entry.place(x=850,y=340)


# 관리자 모드 - 비밀번호 확인 전용
class Settings_Check(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="admin_auth.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)
        
        pw_ent = tk.Entry(self,width=30,show='*')
        pw_ent.place(x=850,y=340)
        
        def pw_auth(n):
            global password
            pw_input = pw_ent.get()
            if pw_input == password:
                master.switch_frame(Settings)
            else:
                msgbox.showwarning("경고!", "잘못된 비밀번호 입력입니다!")
                master.switch_frame(Home)

        pw_ent.bind("<Return>",pw_auth)

# 관리자 모드 승인(완)
class Settings(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="admined.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Home))
        home_button.place(x=30,y=30)
        self.admin_btn1 = tk.PhotoImage(file='admin_btn1.png')
        self.admin_btn2 = tk.PhotoImage(file='admin_btn2.png')
        self.admin_btn3 = tk.PhotoImage(file='admin_btn3.png')

        btn1 = tk.Button(self,width=800,height=100,image=self.admin_btn1,command=lambda:master.switch_frame(Settings_1))
        btn2 = tk.Button(self,width=800,height=100,image=self.admin_btn2,command=lambda:master.switch_frame(Settings_2))
        btn3 = tk.Button(self,width=800,height=100,image=self.admin_btn3,command=lambda:master.switch_frame(Settings_3))

        btn1.place(x=560,y=350)
        btn2.place(x=560,y=520)
        btn3.place(x=560,y=690)

# 관리자 모드 1. 수공구 신규 등록 -> 등록의 경우는 *처리 없이 진행
class Settings_1(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        initEntercheck()
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="admin_win1.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Settings))
        home_button.place(x=30,y=30)
        
        newTL_ent = tk.Entry(self,width=30,font=master.title_font)

        newTL_ent.place(x=730,y=380)

        def Register_newTL(newTL_uid,newTL_name,newTL_type):
            if len(newTL_uid) != 0 and len(newTL_name) != 0 and len(newTL_type) !=0:
                global WS_TL_COUNTER
                if findUID(WS_TL,WS_TL_COUNTER,3,newTL_name) != -1:
                    msgbox.showwarning("경고!","이미 등록된 수공구명입니다")
                else:
                    msg = msgbox.askquestion("수공구 신규 등록","수공구 RFID :  "+newTL_uid+"\n수공구 종류 :  "+newTL_type+"\n수공구 이름 :  "+newTL_name)
                    if msg == 'yes':
                        WS_TL.cell(row=WS_TL_COUNTER,column=1,value=newTL_uid)       
                        WS_TL.cell(row=WS_TL_COUNTER,column=2,value=newTL_type)
                        WS_TL.cell(row=WS_TL_COUNTER,column=3,value=newTL_name)
                        WS_TL.cell(row=WS_TL_COUNTER,column=4,value=0)              # 대여현황 1 : true, 0 : false
                        WS_TL.cell(row=WS_TL_COUNTER,column=5,value="")
                        wb.save(excel_name)
                        WS_TL_COUNTER += 1
                        master.switch_frame(Settings)
            else:
                msgbox.showwarning("경고!","입력한 정보를 확인해주십시오")

        self.comp_btn = tk.PhotoImage(file='완료.png')
        def Check_newTL(n):
            if Entercheck == 0:
                newTL_uid = newTL_ent.get()
                newTL_location = findUID(WS_TL,WS_TL_COUNTER,1,newTL_uid)           # 수공구 DB 대조용
                newTL_location2 = findUID(WS_H,WS_H_COUNTER,1,newTL_uid)            # 사람 DB 대조용
                if newTL_location != -1:
                    newTL_ent.delete(0,'end')
                    msgbox.showwarning("경고!","이미 등록된 수공구입니다")
                elif newTL_location2 != -1:
                    newTL_ent.delete(0,'end')
                    msgbox.showwarning("경고!","이미 사용자로 등록된 TAG입니다")
                else:
                    typelbl = tk.Label(self,text="종류",fg='white',bg='#3B3838',font=master.title_font)
                    newTL_type = tk.Entry(self,width=30,font=master.title_font)
                    namelbl = tk.Label(self,text="수공구명",fg='white',bg='#3B3838',font=master.title_font)
                    newTL_name = tk.Entry(self,width=30,font=master.title_font)
                    
                    newTL_confirm = tk.Button(self,width=160,height=90,image=self.comp_btn,command=lambda:Register_newTL(newTL_uid,newTL_name.get(),newTL_type.get()))
                    countEntercheck()
                    typelbl.place(x=930,y=450)
                    newTL_type.place(x=730,y=500)
                    namelbl.place(x=905,y=570)
                    newTL_name.place(x=730,y=620)
                    newTL_confirm.place(x=880,y=720)

        newTL_ent.bind("<Return>",Check_newTL)
import pyautogui
class blank(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.bg_img = tk.PhotoImage(file="admin_win2.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        ent = tk.Entry(self,bg='#3B3838')
        def enter(n):
            master.switch_frame(Settings_2)
        
        ent.place(x=50,y=50)
        ent.bind("<Return>",enter)
        pyautogui.press('tab')
        pyautogui.press('enter')
        
# 관리자 모드 2. 대여 수공구 현황 확인 및 관리
class Settings_2(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="admin_win2.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Settings))
        home_button.place(x=30,y=30)


        listbox = tk.Listbox(self,width=30,font=master.title_font)
        
        rental_list = []
        # 수공구명 - 사람 이름
        for i in range(2,WS_TL_COUNTER):
            if WS_TL.cell(row=i,column=4).value == 1:
                rental_TL = WS_TL.cell(row=i,column=3).value
                H_location = findUID(WS_H,WS_H_COUNTER,1,WS_TL.cell(row=i,column=5).value)
                rental_name = WS_H.cell(row=H_location,column=2).value
                rental_cell = rental_TL+"_"+rental_name
                rental_list.append(rental_cell)
        
        for i in range(0,len(rental_list)):
            listbox.insert(i,rental_list[i])

        listbox.place(x=720,y=340)
        self.comp_btn = tk.PhotoImage(file='admin_win2btn.png')
        def forceRtn(Rtnchoice,listbox,rental_list):
            if len(rental_list) == 0:
                msgbox.showwarning("경고!","반납할 수공구가 없습니다!")
            else:
                index = Rtnchoice[0]
                Rtnstr = rental_list[index]
                TL_Rtn_index = Rtnstr.find('_')
                TL_Rtn = Rtnstr[0:TL_Rtn_index]

                TL_location = findUID(WS_TL,WS_TL_COUNTER,3,TL_Rtn)
                
                if TL_location == -1:
                    msgbox.showwarning("경고!","엑셀 파일에 문제 발생!")
                else:
                    msg = msgbox.askquestion("강제 반납 처리",TL_Rtn+" 를 반납 처리하시겠습니까?")
                    if msg == 'yes':
                        # LOG 공구 강제 반납 기록
                        global WS_LOG_COUNTER
                        TL_rental_uid = WS_TL.cell(row=TL_location,column=5).value
                        H_location = findUID(WS_H,WS_H_COUNTER,1,TL_rental_uid)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=1,value=CALL_TIME())
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=2,value=TL_rental_uid)
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=3,value=WS_H.cell(row=H_location,column=2).value)# 이름
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=4,value=WS_H.cell(row=H_location,column=3).value) # 학번
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=5,value=WS_H.cell(row=H_location,column=5).value) # 학과
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=6,value="강제반납")
                        WS_LOG.cell(row=WS_LOG_COUNTER,column=7,value=WS_TL.cell(row=TL_location,column=3).value)
                        WS_TL.cell(row=TL_location,column=4,value=0)
                        WS_TL.cell(row=TL_location,column=5,value="")
                        WS_LOG_COUNTER += 1 
                        # DATA 공구 강제 반납 기록
                        WS_TL.cell(row=TL_location,column=4,value=0)
                        WS_TL.cell(row=TL_location,column=5,value = '')

                        wb.save(excel_name)
                        wb2.save(excel_log)  
                        master.switch_frame(blank)
                        
                
        forceRtn_btn = tk.Button(self,image=self.comp_btn,width=400,height=150,command=lambda:forceRtn(listbox.curselection(),listbox,rental_list))
        forceRtn_btn.place(x=760,y=800)

# 관리자 모드 3. 신규 등록 인원 확인(프로그램 종료 전까지 신규 등록 인원만 확인 가능)
class Settings_3(tk.Frame):

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # 배경 + 메인 문구
        self.bg_img = tk.PhotoImage(file="admin_win3.png")
        sky = tk.Label(self,image=self.bg_img)
        sky.pack(fill="both")
        # Home으로 가는 버튼 생성   
        home_button = tk.Button(self, width= 100, height= 100, bd=0,image=master.home_btn , command=lambda: master.switch_frame(Settings))
        home_button.place(x=30,y=30)

        listbox = tk.Listbox(self,selectmode='extended',width=70,font=master.title_font)
        for i in range(0,NewH_counter):
            listbox.insert(i,CALL_H(WS_H_COUNTER-1-i))
        listbox.place(x=400,y=350)

        self.comp_btn = tk.PhotoImage(file='admin_win3btn.png')
        def NewConfirmed():
            msg = msgbox.askquestion("신규 등록 인원 확인","신규 등록 인원 확인처리하시겠습니까?")
            if msg == 'yes':
                global NewH_counter
                NewH_counter=0
                master.switch_frame(Settings)
        new_confirm = tk.Button(self,width=400,height=150,image=self.comp_btn,command=lambda:NewConfirmed())
        new_confirm.place(x=760,y=800)

# 직접 실행시켰을때만 실행되길 원하는 코드들만 넣어주는 것
if __name__ == "__main__":
    app = MainStream()
    app.mainloop()